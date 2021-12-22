from django.shortcuts import redirect, render
from .forms import *
from .models import *
import os
import csv
from openpyxl import load_workbook, Workbook, workbook
from fpdf import FPDF
import math
from datetime import date
today = date.today()


def value_of_grade(grade, mapping_of_grade):  # finding value of particular grade
    s = grade.strip()
    if s in mapping_of_grade:
        return mapping_of_grade[s]
    s = s[:-1]
    return mapping_of_grade[s]

def ranges(request):
    v = RangeInput.objects.all()
    left = "1901EE45"
    right = "1901EE44"
    for p in v:
        left = p.left
        right = p.right
        stamp = p.stamp
        sign = p.sign
    left=left.upper()
    right=right.upper()
    left_num = int(left[6] + left[7])
    right_num = int(right[6] + right[7])
    rest=""
    for i in range(0, 6):
        rest+=left[i]
    if os.path.exists(r"transcriptsIITP") == False:
        os.mkdir(r"transcriptsIITP/")  # folder where output will be stored
    roll_to_name = {}
    subject = {}
    credits_taken = {}
    numerator = {}
    branch = {
        "CS": "Computer Science and Engineering",
        "EE": "Electrical Engineering",
        "ME": "Mechanical Engineering",
        "CE": "Civil Engineering",
        "MM": "Metallurgical and Material Science Engineering",
        "CB": "Chemical and Biochemical Engineering",
        "MA": "Maths",
        "PH": "Physics",
        "CH": "Chemistry"
    }
    degree = {
        "01": "B.Tech",
        "11": "M.Tech",
        "12": "M.Sc",
        "21": "Phd"
    }
    dict_grades = {
        "AA": 10,
        "AB": 9,
        "BB": 8,
        " BB": 8,
        "BC": 7,
        "CC": 6,
        "CD": 5,
        "DD": 4,
        "F": 0,
        "I": 0,
        "AA*": 10,
        "AB*": 9,
        "BB*": 8,
        "BC*": 7,
        "CC*": 6,
        "CD*": 5,
        "DD*": 4,
        "F*": 0,
        "I*": 0,
    }
    pop_list=[]

    with open('media/file/subjects_master.csv', 'r') as curr_file:  # mapping subject code with subject name
        reader = csv.reader(curr_file)
        next(reader)
        for col in reader:
            subject_number = col[0]
            name_of_subject = col[1]
            LTP_of_subject = col[2]
            credits_of_subject = col[3]
            subject[subject_number] = [name_of_subject, LTP_of_subject, credits_of_subject]

    serial_number = {}
    with open('media/file/grades.csv', 'r') as curr_file:  # create and update file sem wise
        reader = csv.reader(curr_file)
        next(reader)
        for col in reader:
            roll = col[0]
            sem_number = col[1]
            subject_code = col[2]
            credit = col[3]
            grade = col[4]
            tup = (roll, sem_number)
            lis = [subject_code, subject[subject_code][0], subject[subject_code][1], credit, grade]
            heading = ["SubCode", "Subject Name", "LTP", "Crd", "Grd"]
            if tup not in serial_number:
                serial_number[tup] = [heading]
            serial_number[tup].append(lis)

    for last_two in range(left_num, right_num+1):
        add=str(last_two)
        if last_two<10:
            add="0"+add
        roll=rest+str(add)
        f = 0
        for i in range(1, 9):
            if (roll, str(i)) in serial_number:
                f = 1
        print((roll, f))
        if f == 0:
            pop_list.append(roll)
            continue


    with open('media/file/names-roll.csv', 'r') as curr_file:  # mapping roll number with name
        reader = csv.reader(curr_file)
        next(reader)
        for col in reader:
            roll = col[0]
            num = int(roll[6] + roll[7])
            curr_rest=""
            for i in range(0, 6):
                curr_rest+=roll[i]
            if num<left_num or num>right_num or curr_rest!=rest:
                continue
            student_name = col[1]
            br = roll[4] + roll[5]
            Course = branch[br]
            year = "20" + roll[0] + roll[1]
            type_degree = roll[2] + roll[3]
            programme = ""
            if type_degree not in degree:
                programme = "Phd"
            else:
                programme = degree[type_degree]
            if type_degree == "01":
                pdf = FPDF(orientation='L', unit='mm', format='A4')
            else:
                pdf = FPDF(orientation='L', unit='mm', format='A3')
            pdf.add_page()
            pdf.set_font("Times", size=10)
            line_height = pdf.font_size * 2.5
            image = r"media/image/heading.png"
            sizew = pdf.epw * 1.05
            sizeh = 50
            pdf.image(image, x=0, y=0, w=sizew, h=sizeh)
            listt = [["Roll: " + roll, "Name: " + student_name, "Year: " + year],
                     ["Programme: " + programme, "Course: " + Course, " "]]
            pdf.ln()
            pdf.set_y(45)
            cnt = 0
            gap = -5
            for row in listt:
                if cnt:
                    pdf.set_y(pdf.get_y() + gap)
                pdf.set_y(pdf.get_y() + 5)
                pdf.set_x(pdf.get_x() + 69)
                for datum in row:
                    if datum != row[1]:
                        pdf.multi_cell(31, line_height, datum, border=1, ln=3, max_line_height=pdf.font_size)
                    else:
                        pdf.multi_cell(50, line_height, datum, border=1, ln=3, max_line_height=pdf.font_size)
                pdf.ln(line_height)
                cnt += 1

            # what up
            pdf.set_y(pdf.get_y() + 5)
            pdf.set_x(pdf.get_x() - 2)
            y = pdf.get_y()
            pdf.set_font("Times", size=5)
            line_height = pdf.font_size * 1.5
            cpi_num = 0
            cpi_den = 0
            for sem in range(1, 5):
                spi_num = 0
                spi_den = 0
                pdf.set_x(pdf.get_x() + (sem - 1) * 69)
                pdf.set_font("Times", size=10)
                pdf.multi_cell(31, line_height, "Semester " + str(sem), border=0, ln=3, max_line_height=pdf.font_size)
                pdf.set_font("Times", size=5)
                pdf.set_y(pdf.get_y() + 5)
                pdf.set_x(pdf.get_x() + (sem - 1) * 69)
                x = pdf.get_x()
                spi_num = 0
                spi_den = 0
                cc = 0
                cnt = 0
                gap = -5
                if (roll, str(sem)) not in serial_number:
                    continue
                for items in serial_number[(roll, str(sem))]:
                    if items != heading:
                        spi_den += int(items[3])
                        spi_num += int(items[3]) * dict_grades[items[4]]
                        cpi_den += int(items[3])
                        cpi_num += int(items[3]) * dict_grades[items[4]]
                        if dict_grades[items[4]]:
                            cc += int(items[3])
                    for datum in items:
                        if datum == items[1]:
                            pdf.multi_cell(35, line_height, datum, border=1, ln=3, max_line_height=pdf.font_size)
                        elif datum == items[0]:
                            pdf.multi_cell(10, line_height, datum, border=1, ln=3, max_line_height=pdf.font_size)
                        elif datum == items[2]:
                            pdf.multi_cell(6, line_height, datum, border=1, ln=3, max_line_height=pdf.font_size)
                        elif datum == items[3]:
                            pdf.multi_cell(5, line_height, datum, border=1, ln=3, max_line_height=pdf.font_size)
                        else:
                            pdf.multi_cell(5, line_height, datum, border=1, ln=3, max_line_height=pdf.font_size)
                    pdf.ln(line_height)
                    pdf.set_x(x)
                # credits_taken-cpi_den
                # credits_cleared-cc
                # spi-spi_num/spi_den
                # cpi-cpi_num/cpi_den
                pdf.ln(line_height)
                pdf.set_x(pdf.get_x() + (sem - 1) * 69)
                s = "Credits Taken: " + str(spi_den) + " Credits Cleared: " + str(cc) + " SPI: " + str(
                    round(spi_num / spi_den, 2)) + " CPI: " + str(round(cpi_num / cpi_den, 2))
                pdf.cell(50, 5, txt=s, ln=1, align='C', border=1)
                pdf.ln(line_height)
                if sem != 4:
                    pdf.set_y(y)

            pdf.ln(line_height)
            y = pdf.get_y()
            pdf.set_font("Times", size=5)
            line_height = pdf.font_size * 1.5
            for sem in range(5, 9):
                pdf.set_x(pdf.get_x() + (sem - 5) * 69)
                pdf.set_font("Times", size=10)
                pdf.multi_cell(31, line_height, "Semester " + str(sem), border=0, ln=3, max_line_height=pdf.font_size)
                pdf.set_font("Times", size=5)
                pdf.set_y(pdf.get_y() + 5)
                pdf.set_x(pdf.get_x() + (sem - 5) * 69)
                x = pdf.get_x()
                spi_num = 0
                spi_den = 0
                cc = 0
                cnt = 0
                gap = -5
                if (roll, str(sem)) not in serial_number:
                    continue
                for items in serial_number[(roll, str(sem))]:
                    if items != heading:
                        spi_den += int(items[3])
                        spi_num += int(items[3]) * dict_grades[items[4]]
                        cpi_den += int(items[3])
                        cpi_num += int(items[3]) * dict_grades[items[4]]
                        if dict_grades[items[4]]:
                            cc += int(items[3])
                    for datum in items:
                        if datum == items[1]:
                            pdf.multi_cell(35, line_height, datum, border=1, ln=3, max_line_height=pdf.font_size)
                        elif datum == items[0]:
                            pdf.multi_cell(10, line_height, datum, border=1, ln=3, max_line_height=pdf.font_size)
                        elif datum == items[2]:
                            pdf.multi_cell(6, line_height, datum, border=1, ln=3, max_line_height=pdf.font_size)
                        elif datum == items[3]:
                            pdf.multi_cell(5, line_height, datum, border=1, ln=3, max_line_height=pdf.font_size)
                        else:
                            pdf.multi_cell(5, line_height, datum, border=1, ln=3, max_line_height=pdf.font_size)
                    pdf.ln(line_height)
                    pdf.set_x(x)
                pdf.ln(line_height)
                pdf.set_x(pdf.get_x() + (sem - 5) * 69)
                s = "Credits Taken: " + str(spi_den) + " Credits Cleared: " + str(cc) + " SPI: " + str(
                    round(spi_num / spi_den, 2)) + " CPI: " + str(round(cpi_num / cpi_den, 2))
                pdf.cell(50, 5, txt=s, ln=1, align='C', border=1)
                pdf.ln(line_height)
                if sem != 4:
                    pdf.set_y(y)
            image = r"media/image/stamp.png"
            sizew = pdf.epw * 0.1
            sizeh = 20
            pdf.image(stamp, x=120, y=169, w=sizew, h=sizeh)

            image = r"media/image/sign.png"
            sizew = pdf.epw * 0.1
            sizeh = 10
            pdf.image(sign, x=240, y=169, w=sizew, h=sizeh)

            image = r"media/image/assistant.jpg"
            sizew = pdf.epw * 0.1
            sizeh = 2
            pdf.image(image, x=240, y=180, w=sizew, h=sizeh)
            # pdf.ln(line_height)

            date = today.strftime("%b-%d-%Y")
            date = "Date of Issue: " + str(date)
            pdf.set_font("Times", size=10)
            pdf.cell(37, 75, txt=date, ln=1, align='C')
            pdf.output('transcriptsIITP/' + roll.upper() + '.pdf')

    if request.method == 'POST' :
        form = TranscriptForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            if request.POST.get('all'):
                return redirect('Proj2')
            if request.POST.get('range'):
                return redirect('range')
    else :
        form = TranscriptForm()
    print(pop_list)
    return render(request, 'polls/HomePage.html', {'form': form, 'pop_list':pop_list, 'f':len(pop_list)})

def solve(request):
    v = RangeInput.objects.all()
    left = "1901EE45"
    right = "1901EE44"
    for p in v:
        left = p.left
        right = p.right
        stamp=p.stamp
        sign=p.sign
    if os.path.exists(r"transcriptsIITP") == False:
        os.mkdir(r"transcriptsIITP/")  # folder where output will be stored
    roll_to_name = {}
    subject = {}
    credits_taken = {}
    numerator = {}
    branch = {
        "CS": "Computer Science and Engineering",
        "EE": "Electrical Engineering",
        "ME": "Mechanical Engineering",
        "CE": "Civil Engineering",
        "MM": "Metallurgical and Material Science Engineering",
        "CB": "Chemical and Biochemical Engineering",
        "MA": "Maths",
        "PH": "Physics",
        "CH": "Chemistry"
    }
    degree = {
        "01": "B.Tech",
        "11": "M.Tech",
        "12": "M.Sc",
        "21": "Phd"
    }
    dict_grades = {
        "AA": 10,
        "AB": 9,
        "BB": 8,
        " BB": 8,
        "BC": 7,
        "CC": 6,
        "CD": 5,
        "DD": 4,
        "F": 0,
        "I": 0,
        "AA*": 10,
        "AB*": 9,
        "BB*": 8,
        "BC*": 7,
        "CC*": 6,
        "CD*": 5,
        "DD*": 4,
        "F*": 0,
        "I*": 0,
    }

    with open('media/file/subjects_master.csv', 'r') as curr_file:  # mapping subject code with subject name
        reader = csv.reader(curr_file)
        next(reader)
        for col in reader:
            subject_number = col[0]
            name_of_subject = col[1]
            LTP_of_subject = col[2]
            credits_of_subject = col[3]
            subject[subject_number] = [name_of_subject, LTP_of_subject, credits_of_subject]

    serial_number = {}
    with open('media/file/grades.csv', 'r') as curr_file:  # create and update file sem wise
        reader = csv.reader(curr_file)
        next(reader)
        for col in reader:
            roll = col[0]
            sem_number = col[1]
            subject_code = col[2]
            credit = col[3]
            grade = col[4]
            tup = (roll, sem_number)
            lis = [subject_code, subject[subject_code][0], subject[subject_code][1], credit, grade]
            heading = ["SubCode", "Subject Name", "LTP", "Crd", "Grd"]
            if tup not in serial_number:
                serial_number[tup] = [heading]
            serial_number[tup].append(lis)

    with open('media/file/names-roll.csv', 'r') as curr_file:  # mapping roll number with name
        reader = csv.reader(curr_file)
        next(reader)
        for col in reader:
            roll = col[0]
            student_name = col[1]
            br = roll[4] + roll[5]
            Course = branch[br]
            year = "20" + roll[0] + roll[1]
            type_degree = roll[2] + roll[3]
            programme = ""
            if type_degree not in degree:
                programme = "Phd"
            else:
                programme = degree[type_degree]
            if type_degree == "01":
                pdf = FPDF(orientation='L', unit='mm', format='A4')
            else:
                pdf = FPDF(orientation='L', unit='mm', format='A3')
            pdf.add_page()
            pdf.set_font("Times", size=10)
            line_height = pdf.font_size * 2.5
            image = r"media/image/heading.png"
            sizew = pdf.epw * 1.05
            sizeh = 50
            pdf.image(image, x=0, y=0, w=sizew, h=sizeh)
            listt = [["Roll: " + roll, "Name: " + student_name, "Year: " + year],
                     ["Programme: " + programme, "Course: " + Course, " "]]
            pdf.ln()
            pdf.set_y(45)
            cnt = 0
            gap = -5
            for row in listt:
                if cnt:
                    pdf.set_y(pdf.get_y() + gap)
                pdf.set_y(pdf.get_y() + 5)
                pdf.set_x(pdf.get_x() + 69)
                for datum in row:
                    if datum != row[1]:
                        pdf.multi_cell(31, line_height, datum, border=1, ln=3, max_line_height=pdf.font_size)
                    else:
                        pdf.multi_cell(50, line_height, datum, border=1, ln=3, max_line_height=pdf.font_size)
                pdf.ln(line_height)
                cnt += 1

            # what up
            pdf.set_y(pdf.get_y() + 5)
            pdf.set_x(pdf.get_x() - 2)
            y = pdf.get_y()
            pdf.set_font("Times", size=5)
            line_height = pdf.font_size * 1.5
            cpi_num = 0
            cpi_den = 0
            for sem in range(1, 5):
                spi_num = 0
                spi_den = 0
                pdf.set_x(pdf.get_x() + (sem - 1) * 69)
                pdf.set_font("Times", size=10)
                pdf.multi_cell(31, line_height, "Semester " + str(sem), border=0, ln=3, max_line_height=pdf.font_size)
                pdf.set_font("Times", size=5)
                pdf.set_y(pdf.get_y() + 5)
                pdf.set_x(pdf.get_x() + (sem - 1) * 69)
                x = pdf.get_x()
                spi_num = 0
                spi_den = 0
                cc = 0
                cnt = 0
                gap = -5
                if (roll, str(sem)) not in serial_number:
                    continue
                for items in serial_number[(roll, str(sem))]:
                    if items != heading:
                        spi_den += int(items[3])
                        spi_num += int(items[3]) * dict_grades[items[4]]
                        cpi_den += int(items[3])
                        cpi_num += int(items[3]) * dict_grades[items[4]]
                        if dict_grades[items[4]]:
                            cc += int(items[3])
                    for datum in items:
                        if datum == items[1]:
                            pdf.multi_cell(35, line_height, datum, border=1, ln=3, max_line_height=pdf.font_size)
                        elif datum == items[0]:
                            pdf.multi_cell(10, line_height, datum, border=1, ln=3, max_line_height=pdf.font_size)
                        elif datum == items[2]:
                            pdf.multi_cell(6, line_height, datum, border=1, ln=3, max_line_height=pdf.font_size)
                        elif datum == items[3]:
                            pdf.multi_cell(5, line_height, datum, border=1, ln=3, max_line_height=pdf.font_size)
                        else:
                            pdf.multi_cell(5, line_height, datum, border=1, ln=3, max_line_height=pdf.font_size)
                    pdf.ln(line_height)
                    pdf.set_x(x)
                #credits_taken-cpi_den
                #credits_cleared-cc
                #spi-spi_num/spi_den
                #cpi-cpi_num/cpi_den
                pdf.ln(line_height)
                pdf.set_x(pdf.get_x() + (sem - 1) * 69)
                s="Credits Taken: "+str(spi_den)+" Credits Cleared: "+str(cc)+" SPI: "+str(round(spi_num/spi_den, 2))+" CPI: "+str(round(cpi_num/cpi_den, 2))
                pdf.cell(50, 5, txt=s,ln=1, align='C', border=1)
                pdf.ln(line_height)
                if sem != 4:
                    pdf.set_y(y)

            pdf.ln(line_height)
            y = pdf.get_y()
            pdf.set_font("Times", size=5)
            line_height = pdf.font_size * 1.5
            for sem in range(5, 9):
                pdf.set_x(pdf.get_x() + (sem - 5) * 69)
                pdf.set_font("Times", size=10)
                pdf.multi_cell(31, line_height, "Semester " + str(sem), border=0, ln=3, max_line_height=pdf.font_size)
                pdf.set_font("Times", size=5)
                pdf.set_y(pdf.get_y() + 5)
                pdf.set_x(pdf.get_x() + (sem - 5) * 69)
                x = pdf.get_x()
                spi_num = 0
                spi_den = 0
                cc = 0
                cnt = 0
                gap = -5
                if (roll, str(sem)) not in serial_number:
                    continue
                for items in serial_number[(roll, str(sem))]:
                    if items != heading:
                        spi_den += int(items[3])
                        spi_num += int(items[3]) * dict_grades[items[4]]
                        cpi_den += int(items[3])
                        cpi_num += int(items[3]) * dict_grades[items[4]]
                        if dict_grades[items[4]]:
                            cc += int(items[3])
                    for datum in items:
                        if datum == items[1]:
                            pdf.multi_cell(35, line_height, datum, border=1, ln=3, max_line_height=pdf.font_size)
                        elif datum == items[0]:
                            pdf.multi_cell(10, line_height, datum, border=1, ln=3, max_line_height=pdf.font_size)
                        elif datum == items[2]:
                            pdf.multi_cell(6, line_height, datum, border=1, ln=3, max_line_height=pdf.font_size)
                        elif datum == items[3]:
                            pdf.multi_cell(5, line_height, datum, border=1, ln=3, max_line_height=pdf.font_size)
                        else:
                            pdf.multi_cell(5, line_height, datum, border=1, ln=3, max_line_height=pdf.font_size)
                    pdf.ln(line_height)
                    pdf.set_x(x)
                pdf.ln(line_height)
                pdf.set_x(pdf.get_x() + (sem - 5) * 69)
                s = "Credits Taken: " + str(spi_den) + " Credits Cleared: " + str(cc) + " SPI: " + str(round(spi_num / spi_den, 2)) + " CPI: " + str(round(cpi_num / cpi_den, 2))
                pdf.cell(50, 5, txt=s, ln=1, align='C', border=1)
                pdf.ln(line_height)
                if sem != 4:
                    pdf.set_y(y)
            image = r"media/image/stamp.png"
            sizew = pdf.epw * 0.1
            sizeh = 20
            pdf.image(stamp, x=120, y=169, w=sizew, h=sizeh)

            image = r"media/image/sign.png"
            sizew = pdf.epw * 0.1
            sizeh = 10
            pdf.image(sign, x=240, y=169, w=sizew, h=sizeh)

            image = r"media/image/assistant.jpg"
            sizew = pdf.epw * 0.1
            sizeh = 2
            pdf.image(image, x=240, y=180, w=sizew, h=sizeh)
            # pdf.ln(line_height)


            date = today.strftime("%b-%d-%Y")
            date = "Date of Issue: "+str(date)
            pdf.set_font("Times", size=10)
            pdf.cell(37 , 75, txt=date, ln=1, align='C')
            pdf.output('transcriptsIITP/' + roll.upper() + '.pdf')

    return redirect('HomePage')