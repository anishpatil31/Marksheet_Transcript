import os
import csv
import sys
from openpyxl import load_workbook, Workbook, workbook
from openpyxl.styles import Font, Fill
from openpyxl.styles import colors
import sys
from .models import *
from django.shortcuts import render, redirect
from .forms import *
import openpyxl
import xlsxwriter
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font


def gcm(request):
    v = Marktsheet_data.objects.all()
    positive = 0
    negative = 0
    for p in v:
        positive = p.positive
        negative = p.negative
    base_dir = r"marksheets"
    if os.path.exists(base_dir) == False:
        os.mkdir(r"marksheets")
    output_path = r"marksheets/concise_marksheet.csv"
    wb = Workbook()
    wb_sheet = wb['Sheet']
    wb_sheet.title = 'concise_marksheet'
    sheet = wb.active
    path = ""
    response = r"media/file/responses.csv"
    heading = ["Timestamp", "Email address", "Google_Score", "Name", "IITP Webmail", "Phone (10 digit only)",
               "Score_After_Negatuve", "Roll_Number"];
    unnamed = "Unnamed: "
    for i in range(7, 35):
        s = unnamed + str(i)
        heading.append(s)
    heading.append("statusAns")
    sheet.append(heading)

    answer_key = []

    flag = "anish"

    with open(os.path.join(path, response), 'r') as responses:
        reader = csv.reader(responses)
        heading_of_columns = next(reader)
        for row in reader:
            if row[6] == "ANSWER":
                for i in range(7, 35):
                    answer_key.append(row[i])
                flag = row[6]
                break

        if flag != "ANSWER":
            if request.method == 'POST':
                form = Markform(request.POST, request.FILES)

                if form.is_valid():
                    form.save()
                    return redirect('Proj1')
            else:
                form = Markform()
            return render(request, 'marktSite/marktSite.html', {'form': form, 'check': 1})

    with open(os.path.join(path, response), 'r') as f:
        read = csv.reader(f)
        h = next(read)
        for rows in read:
            right = 0
            wrong = 0
            not_attempted = 0
            for i in range(0, len(answer_key)):
                if rows[i + 7] == answer_key[i]:
                    right += 1
                elif rows[i + 7] == "":
                    not_attempted += 1
                else:
                    wrong += 1
            total_questions = right + wrong + not_attempted
            l = [rows[0], rows[1]]
            # push google_score here
            google_score = positive * right
            l.append(google_score)
            for i in range(3, 6):
                l.append(rows[i])
            # push score_after_negative here
            tot_score = (positive * right) + (negative * wrong)
            l.append(tot_score)
            for i in range(6, 35):
                l.append(rows[i])
            # push ansstatus
            statusAns = "[" + str(right) + ", " + str(wrong) + ", " + str(not_attempted) + "]"
            l.append(statusAns)
            sheet.append(l)

    wb.save(output_path)
    return redirect('Home')
